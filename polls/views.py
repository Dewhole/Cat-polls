import csv, io
from io import BytesIO
from datetime import date

import xlwt
import requests
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.decorators import permission_required
from django.shortcuts import render, redirect
from django.views.generic.base import View
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import load_workbook

from .models import Users, Questions
from .forms import UsersForm


class QuestionsView(ListView):
    def get(self, request):
        questions = Questions.objects.all()

        return render(request, "polls/polls7.html", locals())


class UsersDetailView(View):
    def get(self, request, slug):

        form = UsersForm()

        user = Users.objects.get(url=slug)
        questions = Questions.objects.all()
        if (
            user.rating1 == 1
            or user.rating1 == 2
            or user.rating1 == 3
            or user.rating1 == 4
            or user.rating1 == 5
            or user.rating1 == 6
            or user.rating1 == 7
            or user.rating1 == 8
            or user.rating1 == 9
            or user.rating1 == 10
        ):
            return redirect("/pass")
        userfactureDate = str(user.factureDate)
        userfactureDate2 = userfactureDate[5:7]
        user.factureDate = userfactureDate2
        if user.factureDate == "01":
            user.factureDate = "январе"
        elif user.factureDate == "02":
            user.factureDate = "феврале"
        elif user.factureDate == "03":
            user.factureDate = "марте"
        elif user.factureDate == "04":
            user.factureDate = "апреле"
        elif user.factureDate == "05":
            user.factureDate = "мае"
        elif user.factureDate == "06":
            user.factureDate = "июне"
        elif user.factureDate == "07":
            user.factureDate = "июле"
        elif user.factureDate == "08":
            user.factureDate = "августе"
        elif user.factureDate == "09":
            user.factureDate = "сентябре"
        elif user.factureDate == "10":
            user.factureDate = "октябре"
        elif user.factureDate == "11":
            user.factureDate = "ноябре"
        elif user.factureDate == "12":
            user.factureDate = "декабре"
        x = " ".join(user.description.split()[0:3])
        xx = x[0:-1] + ""

        data = {"form": form, "user": user}
        error = ""
        usertitle = str(user.title)

        userloaddate = user.loadDate

        now = date.today()

        period = userloaddate - now

        if period.days < -5:
            return redirect("/success")

        if usertitle == "VT Parts Transactional Survey":
            user.description = xx
            return render(request, "polls/1pollsParts.html", data)
        elif usertitle == "VT Field Service Transactional Survey":
            user.description = user.description
            return render(request, "polls/2pollsField.html", data)
        elif usertitle == "VT Shop Service Transactional Survey":
            user.description = user.description
            return render(request, "polls/3pollsShop.html", data)
        elif usertitle == "VT Sales Transactional Survey":
            user.description = user.description
            return render(request, "polls/4pollsSales.html", data)
        elif usertitle == "VT Rental Transactional Survey":
            user.description = user.description
            return render(request, "polls/5pollsRental.html", data)
        elif usertitle == "VT EMPR Parts Transactional Survey":
            user.description = xx
            return render(request, "polls/6EMPRParts.html", data)
        elif usertitle == "VT EMPR Field Service Transactional Survey":
            user.description = user.description
            return render(request, "polls/7EMPRField.html", data)
        else:
            return render(request, "polls/polls.html", data)


def passs(request):
    template = "pass.html"

    prompt = {"order": ""}

    if request.method == "GET":
        return render(request, template, prompt)


def success(request):
    template = "pass.html"

    prompt = {"order": ""}

    if request.method == "GET":
        return render(request, template, prompt)

    # Блок отправки сообщения с файлом на почту


def send_email(addr_to, msg_subj, msg_text):
    addr_from = "VT-Survey@vost-tech.ru"  # Отправитель
    password = "sQi4lO$W"  # Пароль
    msg = MIMEMultipart()  # Создаем сообщение
    msg["From"] = addr_from  # Адресат
    msg["To"] = addr_to  # Получатель
    msg["Subject"] = msg_subj  # Тема сообщения

    body = msg_text  # Текст сообщения
    msg.attach(MIMEText(body, "plain"))  # Добавляем в сообщение текст

    # """\
    # <html>
    # <head></head>
    # <body>
    #    <p>
    #            <i>
    #    ООО «Восточная Техника»
    #    </i><br>
    #    <i>
    #    Tel.:     +7 383 212-56-11
    #    </i><br>
    ##    <i><br>
    #   Web: www.vost-tech.ru  <a href="www.vost-tech.ru">www.vost-tech.ru</a>
    #   </i><br>
    #  <i>
    # Follow us on  <a href="https://www.facebook.com/vosttech/">Facebook</a> | <a href="https://www.youtube.com/channel/UCnOaq6Wy4h7sjjQKkcRINPQ">Instagram</a> | <a href="https://www.instagram.com/vosttech/">YouTube</a>
    # </i>
    # </p>
    # </body>
    # </html>
    ###"""
    # ======== Этот блок настраивается для каждого почтового провайдера отдельно ===============================================
    server = smtplib.SMTP("smtp.office365.com", 587)  # Создаем объект SMTP
    server.starttls()  # Начинаем шифрованный обмен по TLS
    server.login(addr_from, password)  # Получаем доступ
    server.send_message(msg)  # Отправляем сообщение
    server.quit()  # Выходим
    # ==========================================================================================================================


@permission_required("admin_can_add_log_entry", login_url="/admin/")
def xls(request):
    template = "contact_upload_xls.html"

    prompt = {"order": ""}

    if request.method == "GET":
        return render(request, template, prompt)

    try:
        xls_file = request.FILES["file"].read()
        wb = load_workbook(filename=BytesIO(xls_file))

        sheet = wb.active

        # print(sheet[2][0].value)

        for column in range(2, sheet.max_row + 1):

            userfacture = str(sheet[column][26].value)
            try:
                entry = Users.objects.get(facture=userfacture)
                continue

            except:
                username = str(sheet[column][0].value)

                userfacture = str(sheet[column][26].value)
                usermail = (str(sheet[column][31].value),)
                userphone = (str(sheet[column][28].value),)
                userphone2 = (str(sheet[column][29].value),)

                usermail2 = "".join(usermail)
                msg_text = (
                    "Уважаемый (ая) "
                    + username
                    + ", благодарим Вас за обращение в «Восточную Технику» и просим принять участие в коротком опросе по ссылке http://survey.vost-tech.ru/"
                    + userfacture
                    + ". Опрос займёт около 5 минут."
                    + "\n \n ООО «Восточная Техника» \n Tel.:     +7 383 212-56-11 \n "
                    + "Web: www.vost-tech.ru  "
                )

                msg_text_phone = (
                    "Уважаемый (ая) "
                    + username
                    + ", благодарим Вас за обращение в «Восточную Технику» и просим принять участие в коротком опросе по ссылке http://survey.vost-tech.ru/"
                    + userfacture
                    + ". Опрос займёт около 5 минут."
                )

                addr_to = usermail2
                if username == "None":
                    break
                else:
                    try:
                        send_email(addr_to, "Оцените нашу работу", msg_text)
                    except:
                        h = 1
                    if username == "None":
                        break
                    else:
                        HEADERS = {
                            "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:82.0) Gecko/20100101 Firefox/82.0",
                            "accept": "*/*",
                        }
                        POST = {
                            "user": "1649851.6",
                            "pass": "U$fSQs4u8iMx",
                            "action": "post_sms",
                            "sender": "VOST_TECH",
                            "message": msg_text_phone,
                            "target": userphone,
                        }
                        POST2 = {
                            "user": "1649851.6",
                            "pass": "U$fSQs4u8iMx",
                            "action": "post_sms",
                            "sender": "VOST_TECH",
                            "message": msg_text_phone,
                            "target": userphone2,
                        }
                        s = requests.get(
                            "https://beeline.amega-inform.ru/sms_send/",
                            headers=HEADERS,
                            params=POST,
                        )
                        ss = requests.get(
                            "https://beeline.amega-inform.ru/sms_send/",
                            headers=HEADERS,
                            params=POST2,
                        )

        for column in range(2, sheet.max_row + 1):
            userfacture = str(sheet[column][26].value)

            try:
                entry = Users.objects.get(facture=userfacture)
                continue

            except:
                theme = sheet[column][32].value
                if theme == "VT Parts Transactional Survey":
                    _, created = Users.objects.update_or_create(
                        name=sheet[column][0].value,
                        partyID=sheet[column][3].value,
                        company=sheet[column][4].value,
                        description=sheet[column][25].value,
                        facture=sheet[column][26].value,
                        factureDate=sheet[column][27].value,
                        phone=sheet[column][28].value,
                        phoneCompany=sheet[column][29].value,
                        phoneMobile=sheet[column][30].value,
                        mail=sheet[column][31].value,
                        url=sheet[column][26].value,
                        title=Questions.objects.get(pk=11),
                    )

                elif theme == "VT Field Service Transactional Survey":
                    _, created = Users.objects.update_or_create(
                        name=sheet[column][0].value,
                        partyID=sheet[column][3].value,
                        company=sheet[column][4].value,
                        description=sheet[column][25].value,
                        facture=sheet[column][26].value,
                        factureDate=sheet[column][27].value,
                        phone=sheet[column][28].value,
                        phoneCompany=sheet[column][29].value,
                        phoneMobile=sheet[column][30].value,
                        mail=sheet[column][31].value,
                        url=sheet[column][26].value,
                        title=Questions.objects.get(pk=12),
                    )
                elif theme == "VT Shop Service Transactional Survey":
                    _, created = Users.objects.update_or_create(
                        name=sheet[column][0].value,
                        partyID=sheet[column][3].value,
                        company=sheet[column][4].value,
                        description=sheet[column][25].value,
                        facture=sheet[column][26].value,
                        factureDate=sheet[column][27].value,
                        phone=sheet[column][28].value,
                        phoneCompany=sheet[column][29].value,
                        phoneMobile=sheet[column][30].value,
                        mail=sheet[column][31].value,
                        url=sheet[column][26].value,
                        title=Questions.objects.get(pk=13),
                    )
                elif theme == "VT Sales Transactional Survey":
                    _, created = Users.objects.update_or_create(
                        name=sheet[column][0].value,
                        partyID=sheet[column][3].value,
                        company=sheet[column][4].value,
                        description=sheet[column][25].value,
                        facture=sheet[column][26].value,
                        factureDate=sheet[column][27].value,
                        phone=sheet[column][28].value,
                        phoneCompany=sheet[column][29].value,
                        phoneMobile=sheet[column][30].value,
                        mail=sheet[column][31].value,
                        url=sheet[column][26].value,
                        title=Questions.objects.get(pk=14),
                    )
                elif theme == "VT Rental Transactional Survey":
                    _, created = Users.objects.update_or_create(
                        name=sheet[column][0].value,
                        partyID=sheet[column][3].value,
                        company=sheet[column][4].value,
                        description=sheet[column][25].value,
                        facture=sheet[column][26].value,
                        factureDate=sheet[column][27].value,
                        phone=sheet[column][28].value,
                        phoneCompany=sheet[column][29].value,
                        phoneMobile=sheet[column][30].value,
                        mail=sheet[column][31].value,
                        url=sheet[column][26].value,
                        title=Questions.objects.get(pk=15),
                    )
                elif theme == "VT EMPR Parts Transactional Survey":
                    _, created = Users.objects.update_or_create(
                        name=sheet[column][0].value,
                        partyID=sheet[column][3].value,
                        company=sheet[column][4].value,
                        description=sheet[column][25].value,
                        facture=sheet[column][26].value,
                        factureDate=sheet[column][27].value,
                        phone=sheet[column][28].value,
                        phoneCompany=sheet[column][29].value,
                        phoneMobile=sheet[column][30].value,
                        mail=sheet[column][31].value,
                        url=sheet[column][26].value,
                        title=Questions.objects.get(pk=16),
                    )
                elif theme == "VT EMPR Field Service Transactional Survey":
                    _, created = Users.objects.update_or_create(
                        name=sheet[column][0].value,
                        partyID=sheet[column][3].value,
                        company=sheet[column][4].value,
                        description=sheet[column][25].value,
                        facture=sheet[column][26].value,
                        factureDate=sheet[column][27].value,
                        phone=sheet[column][28].value,
                        phoneCompany=sheet[column][29].value,
                        phoneMobile=sheet[column][30].value,
                        mail=sheet[column][31].value,
                        url=sheet[column][26].value,
                        title=Questions.objects.get(pk=17),
                    )

                context = {}

        messages.info(request, "Файл успешно импортирован!")

        return redirect("/admin")
    except:
        return redirect("/xls")


@permission_required("admin_can_add_log_entry", login_url="/admin/")
def contact_upload(request):
    template = "contact_upload.html"

    prompt = {"order": ""}

    if request.method == "GET":
        return render(request, template, prompt)

    csv_file = request.FILES["file"]

    if not csv_file.name.endswith(".csv"):
        messages.error(request, "This is not aa csv file")

    data_set = csv_file.read().decode("UTF-8")
    io_string = io.StringIO(data_set)
    io_string2 = io.StringIO(data_set)
    next(io_string)
    next(io_string2)
    # Считывание данных из CSV файла
    for column in csv.reader(io_string2, delimiter=";", quotechar="|"):
        username = column[0]
        userfacture = column[26]
        # msg_text = "Здравствуйте уважаемый, " + username + ", предлагаем пройти опрос по ссылке http://127.0.0.1:8000/" + userfacture
        # addr_to = "dewhole1@gmail.com"
        # send_email(addr_to, "Опрос VT", msg_text)

    for column in csv.reader(io_string, delimiter=";", quotechar="|"):

        if column[32] == "1":
            primaryKey = (Questions.objects.get(pk=1),)
            print(primaryKey)
        elif column[32] == "2":
            primaryKey = (Questions.objects.get(pk=2),)
            print(primaryKey)
        elif column[32] == 3:
            primaryKey = (Questions.objects.get(pk=3),)
        elif column[32] == 4:
            primaryKey = (Questions.objects.get(pk=4),)
        elif column[32] == 5:
            primaryKey = (Questions.objects.get(pk=5),)
        elif column[32] == 6:
            primaryKey = (Questions.objects.get(pk=6),)

        _, created = Users.objects.update_or_create(
            name=column[0],
            partyID=column[3],
            company=column[4],
            description=column[25],
            facture=column[26],
            factureDate=column[27],
            phone=column[28],
            phoneCompany=column[29],
            phoneMobile=column[30],
            mail=column[31],
            url=column[26],
            title=primaryKey,
        )
        context = {}
    messages.info(request, "Файл успешно импортирован!")

    return render(request, template, context)


@permission_required("admin_can_add_log_entry", login_url="/admin/")
def export(request):

    if request.method == "GET":

        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = 'attachment; filename="users.xls"'

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Users")

        # Sheet header, first row
        row_num = 0

        font_style = xlwt.XFStyle()
        font_style.font.bold = True

        columns = [
            "К/л",
            "Анкета",
            "Список обзвона",
            "Код каталога",
            "Компания",
            "Дата звонка",
            "Время звонка",
            "Длительность звонка (в минутах)",
            "Звонок успешен",
            "Ответ1",
            "Примечание1",
            "Ответ2",
            "Примечание2",
            "Ответ3",
            "Примечание3",
            "Ответ4",
            "Примечание4",
            "Ответ5",
            "Примечание5",
            "Ответ6",
            "Примечание6",
            "Ответ7",
            "Примечание7",
            "Ответ8",
            "Примечание8",
            "Ответ9",
            "Примечание9",
            "Ответ10",
            "Примечание10",
            "Ответ11",
            "Примечание11",
            "Ответ12",
            "Примечание12",
            "Описание",
            "фактура",
            "Дата фактуры",
            "Телефон",
            "Телефон организации",
            "Мобильный телефон",
            "Эл. почта",
            "Тема опроса",
        ]

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        rows = Users.objects.all().values_list(
            "name",
            "anketa",
            "listPhone",
            "partyID",
            "company",
            "datePhone",
            "timePhone",
            "timesPhone",
            "lucky",
            "rating1",
            "unswer1",
            "rating2",
            "unswer2",
            "rating3",
            "unswer3",
            "rating4",
            "unswer4",
            "rating5",
            "unswer5",
            "rating6",
            "unswer6",
            "rating7",
            "unswer7",
            "rating8",
            "unswer8",
            "rating9",
            "unswer9",
            "rating10",
            "unswer10",
            "rating11",
            "unswer11",
            "rating12",
            "unswer12",
            "description",
            "facture",
            "factureDate",
            "phone",
            "phoneCompany",
            "phoneMobile",
            "mail",
            "title",
        )
        for row in rows:
            if row[40] == 11:
                lst = list(row)
                lst[40] = "VT Parts Transactional Survey"
                row = tuple(lst)

            elif row[40] == 12:
                lst = list(row)
                lst[40] = "VT Field Service Transactional Survey"
                row = tuple(lst)

            elif row[40] == 13:
                lst = list(row)
                lst[40] = "VT Shop Service Transactional Survey"
                row = tuple(lst)

            elif row[40] == 14:
                lst = list(row)
                lst[40] = "VT Sales Transactional Survey"
                row = tuple(lst)

            elif row[40] == 15:
                lst = list(row)
                lst[40] = "VT Rental Transactional Survey"
                row = tuple(lst)

            elif row[40] == 16:
                lst = list(row)
                lst[40] = "VT EMPR Parts Transactional Survey"
                row = tuple(lst)

            elif row[40] == 17:
                lst = list(row)
                lst[40] = "VT EMPR Field Service Transactional Survey"
                row = tuple(lst)

            else:
                row = row
            row_num += 1
            for col_num in range(len(row)):
                ws.write(row_num, col_num, row[col_num], font_style)

        wb.save(response)
        return response


@permission_required("admin_can_add_log_entry", login_url="/admin/")
def default(request):

    return render(request, "polls/default.html")


class addUnswer(View):
    def post(self, request, pk):

        form = UsersForm(request.POST, instance=Users.objects.get(pk=pk))
        if form.is_valid():

            form.save()
            return redirect("/pass")


""" 

@permission_required('admin_can_add_log_entry', login_url='/admin/')
def export(request):


    if request.method == "GET":
        
        response = HttpResponse(content_type='text/csv', charset='utf-16-le')
        response['Content-Disposition'] = 'attachment; filename=results.csv'
        response.write(codecs.BOM_UTF16_LE)
        writer = csv.writer(response, dialect='excel-tab')

        writer.writerow(['К/л', 'Анкета', 'Список обзвона', 'Код каталога', 'Компания', 'Дата звонка', 'Время звонка',  'Длительность звонка (в минутах)', 'Звонок успешен', 'Ответ1', 'Примечание1', 'Ответ2', 'Примечание2', 'Ответ3', 'Примечание3', 'Ответ4', 'Примечание4', 'Ответ5', 'Примечание5', 'Ответ6', 'Примечание6', 'Ответ7', 'Примечание7', 'Ответ8', 'Примечание8', 'Ответ9', 'Примечание9', 'Ответ10', 'Примечание10', 'Ответ11', 'Примечание11', 'Ответ12', 'Примечание12', 'Описание', 'фактура', 'Дата фактуры', 'Телефон', 'Телефон организации', 'Мобильный телефон', 'Эл. почта', 'Тема опроса' ])

        for user in Users.objects.all().values_list('name', 'anketa', 'listPhone', 'partyID', 'company', 'datePhone', 'timePhone', 'timesPhone', 'lucky', 'rating1', 'unswer1', 'rating2', 'unswer2', 'rating3', 'unswer3', 'rating4', 'unswer4', 'rating5', 'unswer5', 'rating6', 'unswer6', 'rating7', 'unswer7', 'rating8', 'unswer8', 'rating9', 'unswer9', 'rating10', 'unswer10', 'rating11', 'unswer11', 'rating12', 'unswer12', 'description', 'facture', 'factureDate' , 'phone','phoneCompany', 'phoneMobile', 'mail', 'title'):
            if user[40] == 11:
                lst = list(user)
                lst[40] = 'VT Parts Transactional Survey'
                user = tuple(lst)
                writer.writerow(user)
            elif user[40] == 12:
                lst = list(user)
                lst[40] = "VT Field Service Transactional Survey"
                user = tuple(lst)
                writer.writerow(user)
            elif user[40] == 13:
                lst = list(user)
                lst[40] = 'VT Shop Service Transactional Survey'
                user = tuple(lst)
                writer.writerow(user)
            elif user[40] == 14:
                lst = list(user)
                lst[40] = "VT Sales Transactional Survey"
                user = tuple(lst)
                writer.writerow(user)

            elif user[40] == 15:
                lst = list(user)
                lst[40] = "VT Rental Transactional Survey"
                user = tuple(lst)
                writer.writerow(user)
            elif user[40] == 16:
                lst = list(user)
                lst[40] = 'VT EMPR Parts Transactional Survey'
                user = tuple(lst)
                writer.writerow(user)
            elif user[40] == 17:
                lst = list(user)
                lst[40] = 'VT EMPR Field Service Transactional Survey'
                user = tuple(lst)
                writer.writerow(user)
            else:
                writer.writerow(user)

        return response

 """
